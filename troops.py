import openpyxl as xl
from member import *
from nav import *

troops = []
just_troops = []
spells = []
siege_troops = []
troop_files = os.listdir("images/troops_new")
troop_directory = 'images/troops_new/'
slide_position = 1

# print("Troop files:", troop_files)

def get_image_from_file(troop, style):
    file = f'{troop}_{style}.png'
    if style == "": file = f'{troop}.png'
    image = None
    if file in troop_files:
        image = cv2.imread(troop_directory + file, 0)
    else:
        print(f"Adding troop image: could not find '{file}'")
    return image

def get_image(troop, style):
    object = next((x for x in troops if x.name == troop), None)
    if style == "army": return object.army
    if style == "train": return object.train
    if style == "donate1": return object.donate1
    if style == "donate2": return object.donate2
    if style == "attack": return object.attack
    if style == "research": return object.research
    return

def get_super_troop(troop):
    goto(main)
    hold_key("a", 0.5)
    hold_key("w", 0.5)
    sequence = ["super_boost/boost", "super_boost/boost_on", f"super_boost/{troop.name}", "super_boost/potion", "super_boost/dark", "super_boost/potion_small", "super_boost/dark2", ]
    for image in sequence:
        time.sleep(0.5)
        val, loc, rect = find_cv2(image)
        print(image, val)
        if val > 0.6:
            print("Click")
            click_rect(rect)
    pag.press("esc")


class Troop():
    def __init__(self, name, type, slide, training_time, donate_bool, donate_preference, donations, donation_count):
        self.name = name
        self.type = type
        self.slide = slide
        self.training_time = training_time
        self.donate_bool = donate_bool
        self.donate_preference = donate_preference
        self.currently_training = False
        if type != "hero":
            self.i_army = Image(name=f"i_{name}_army", file=f'{troop_directory}{name}_army.png', no_of_regions=1, type="army", threshold=0.72)
            self.i_train = Image(name=f"i_{name}_train", file=f'{troop_directory}{name}_train.png', no_of_regions=1, type="train")
            self.i_training = Image(name=f"i_{name}_training", file=f'{troop_directory}{name}_training.png', no_of_regions=1, type="training")
            self.i_donate1 = Image(name=f"i_{name}_donate1", file=f'{troop_directory}{name}_donate1.png', no_of_regions=1, type="donate1")
            self.i_donate2 = Image(name=f"i_{name}_donate2", file=f'{troop_directory}{name}_donate2.png', no_of_regions=2, type="donate2")
            if "super" not in name:
                self.i_research = Image(name=f"i_{name}_research", file=f'{troop_directory}{name}_research.png', no_of_regions=1, type="research")
        else:
            self.i_army = None
            self.i_train = None
            self.i_training = None
            self.i_donate1 = None
            self.i_donate2 = None
            self.research = None
        self.i_attack = Image(name=f"i_{name}_attack", file=f'{troop_directory}{name}_attack.png')

        self.donations = donations
        self.donation_count = donation_count
        self.super_troop = False
        if "super" in self.name:
            self.super_troop = True

        troops.append(self)
        if type == "troop": just_troops.append(self)
        if type == "spell": spells.append(self)
        if type == "siege": siege_troops.append(self)

    def __str__(self):
        return self.name

    def in_castle(self):
        val, loc, rect = find(self.i_army.image, get_screenshot(CASTLE_TROOPS), show_image=False)
        return val > 0.8

    def start_train(self, count, move_to_start=False):
        global slide_position
        print("Start train:", self.name, self.type, count)
        # Set up
        if self.type == "troop":
            goto(troops_tab)
            slide(slide_position, self.slide)
        if self.type == "spell":
            goto(spells_tab)
        if self.type == "siege":
            goto(siege_tab)

        val, loc, rect = find(self.i_train.image, get_screenshot(TRAIN_RANGE))
        if val < 0.8 and self.super_troop:
            get_super_troop(self)
            goto(troops_tab)

        if val > 0.8:
            for x in range(count):
                click_rect(rect, region=TRAIN_RANGE)
        # val, outcome = click(self.train, region=TRAIN_RANGE, show_image=False)
        print("Create troop:", self.name, round(val,2))
        if move_to_start:
            move_to_queue_start(self)

        print("Start train:", self.name, count)
        return self.training_time * count

    def delete(self, count):
        print("Troop delete")
        if self.i_army.image is None: return
        goto(army_tab)
        click_cv2("edit_army")
        val, loc, rect = find(self.i_army.image, get_screenshot(ARMY_EXISTING))
        rect_adj = [rect[0] + ARMY_EXISTING[0], rect[1] + ARMY_EXISTING[1], rect[2], rect[3]]
        spot = pag.center(rect_adj)
        for x in range(count):
            pag.click(spot)
        click_cv2("edit_army_okay")
        click_cv2("surrender_okay")

levels_filename = 'C:/Users/darre/OneDrive/Darren/clash_bot/levels.xlsx'

def load_troops():
    wb = xl.load_workbook(levels_filename)
    sheet = wb['Troops']
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, 1).value
        type = sheet.cell(row, 2).value
        slide = sheet.cell(row, 3).value
        training_time = sheet.cell(row, 4).value
        donate_bool = sheet.cell(row, 5).value
        donate_preference = sheet.cell(row, 6).value
        donations = sheet.cell(row, 7).value
        donation_count = sheet.cell(row, 8).value
        if name == "goblin":
            donate_bool = True
            donate_preference = 10
            donation_count = 10
            donations = 1

        print("Creating troop:", name)
        Troop(name=name, type=type, slide=slide, training_time=training_time, donate_bool=donate_bool,
              donate_preference=donate_preference, donations=donations, donation_count=donation_count)

    print()

load_troops()

# print()
# for name, type, slide, training_time, donate_bool, donate_preference, donations, donation_count in troop_data:
#     print("Creating troop:", name)
#     Troop(name=name, type=type, slide=slide, training_time=training_time, donate_bool=donate_bool,
#           donate_preference=donate_preference, donations=donations, donation_count=donation_count)

barb = next((x for x in troops if x.name == 'barb'), None)
archer = next((x for x in troops if x.name == 'archer'), None)
goblin = next((x for x in troops if x.name == 'goblin'), None)
giant = next((x for x in troops if x.name == 'giant'), None)
wizard = next((x for x in troops if x.name == 'wizard'), None)
bomber = next((x for x in troops if x.name == 'bomber'), None)
bloon = next((x for x in troops if x.name == 'bloon'), None)
dragon = next((x for x in troops if x.name == 'dragon'), None)
baby_drag = next((x for x in troops if x.name == 'baby_dragon'), None)
edrag = next((x for x in troops if x.name == 'edrag'), None)
minion = next((x for x in troops if x.name == 'minion'), None)
hog = next((x for x in troops if x.name == 'hog'), None)
golem = next((x for x in troops if x.name == 'golem'), None)
witch = next((x for x in troops if x.name == 'witch'), None)
lava_hound = next((x for x in troops if x.name == 'lava_hound'), None)
ice_golem = next((x for x in troops if x.name == 'ice_golem'), None)

super_barb = next((x for x in troops if x.name == 'super_barb'), None)

lightening = next((x for x in troops if x.name == 'lightening'), None)
heal = next((x for x in troops if x.name == 'heal'), None)
rage = next((x for x in troops if x.name == 'rage'), None)
freeze = next((x for x in troops if x.name == 'freeze'), None)
poison = next((x for x in troops if x.name == 'poison'), None)
skeleton = next((x for x in troops if x.name == 'skeleton'), None)
clone = next((x for x in troops if x.name == 'clone'), None)

queen = next((x for x in troops if x.name == 'queen'), None)
king = next((x for x in troops if x.name == 'king'), None)
warden = next((x for x in troops if x.name == 'warden'), None)
champ = next((x for x in troops if x.name == 'champ'), None)

ram = next((x for x in troops if x.name == 'ram'), None)
log_thrower = next((x for x in troops if x.name == 'log_thrower'), None)

troops.sort(key=lambda x: x.donate_preference, reverse=False)

def troop_str(troops):
    string = ""
    for x in troops:
        try:
            string += x.name + ", "
        except:
            pass
    return string[0:-1]


def delete_a_troop():
    val, loc, rect = find_cv2("remove_troops", DELETE_2_REGION)
    center = pag.center(rect)
    pag.click(center)

def make_room(troop):
    count = 0
    colour = check_troop_colour_train(troop)
    print("Make room:", colour)
    while not colour and count < 6:
        delete_a_troop()
        colour = check_troop_colour_train(troop)
        count += 1

def check_troop_colour_train(troop):
    if troop.train is None:
        print("No training image")
        return False
    val, loc, rect = find(troop.train, get_screenshot(TRAIN_RANGE), troop.name)
    rect_adj = [rect[0] + TRAIN_RANGE[0], rect[1] + TRAIN_RANGE[1], rect[2], rect[3], ]
    colour = check_colour_rect(rect_adj, show_image=False)
    return colour

def move_to_queue_start(troop):
    print("Moving to queue start")
    for _ in range(2):
        val, loc, rect = find(troop.training, get_screenshot(BACKLOG))
        if val > 0.70:
            a = BACKLOG[0] + pag.center(rect)[0], BACKLOG[1] + pag.center(rect)[1] + 10
            b = BACKLOG[0] + BACKLOG[2] + 10, a[1] + 10
            if a[0] < 1600: drag(a, b)

def drag(a, b):
    x0, y0 = a
    x1, y1 = b
    pag.moveTo(x0, y0, 0.3)
    pag.dragTo(x1, y1, 1, button="left")
    time.sleep(1)

def slide(slide_pos, slide_pos_target):
    time.sleep(0.2)
    print(f"Slide from {slide_pos} to {slide_pos_target}")
    if slide_pos == slide_pos_target:
        return slide_pos

    if slide_pos < slide_pos_target:
        start_x = 1650
        end_x = 250
        slide_pos = min(slide_pos + 1, 2)
    else:
        start_x = 250
        end_x = 1650
        slide_pos = min(slide_pos - 1, 1)

    dur = 0.3
    pag.moveTo(start_x, 600, dur)
    pag.dragTo(end_x, 600, dur)
    time.sleep(1)
    return slide_pos

def merge_troop_regions():
    for image_type in ["donate1", "donate2"]:
        all_images = [x for x in images if image_type in x.name]
        print("List of images")
        for image in all_images: print(image.name)

        print("Adding regions")
        for image1 in all_images:
            for image2 in all_images:
                if image1 == image2: continue
                for region in image1.regions:
                    if region not in image2.regions:
                        image2.save_region(region)

        print(all_images)
        for image in all_images: print(image.name)

# log_thrower.i_donate2.show_regions()

# for x in troops:
#     print(x)
#     print(x.i_donate2.image)
#     if x.i_donate2.image:
#         show(x.i_donate2)

