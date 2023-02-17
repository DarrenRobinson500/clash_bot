import openpyxl as xl
from tower import *

levels_filename = 'C:/Users/darre/OneDrive/Darren/clash_bot/levels.xlsx'

def load_towers():
    wb = xl.load_workbook(levels_filename)
    sheet = wb['Towers']
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, 1).value.replace(" ", "_").lower()
        village = sheet.cell(row, 2).value.lower()
        # village = "main"
        category = sheet.cell(row, 3).value.lower()
        resource = sheet.cell(row, 4).value.lower()
        priority = sheet.cell(row, 5).value
        if priority is None: priority = 0

        # print("Creating tower:", name)
        Tower(name=name, village=village, category=category, resource=resource, priority=priority)

    print()

def load_levels():
    wb = xl.load_workbook(levels_filename)
    sheet = wb['Levels']
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, 1).value
        number = sheet.cell(row, 2).value
        th = sheet.cell(row, 3).value
        gold = sheet.cell(row, 4).value
        elixir = sheet.cell(row, 5).value
        dark = sheet.cell(row, 6).value
        days = sheet.cell(row, 7).value

        name = name.replace(" ", "_").lower()
        # print(name, level, th, gold, elixir, dark, days)

        if gold is None: gold = 0
        if elixir is None: elixir = 0
        if dark is None: dark = 0

        tower = next((x for x in towers if x.name == name.lower()), None)
        if tower:
            tower.add_level(tower, number, th, gold, elixir, dark, days)
        else:
            print(f"No {name.lower()} found")

def return_tower(tower_name):
    return next((x for x in towers if x.name == tower_name), None)


load_towers()
towers.sort(key=lambda x: x.priority, reverse=False)

load_levels()
# for tower in towers:
#     tower.print_tower()

cannon = return_tower("cannon")
mortar = return_tower("mortar")
wall = return_tower("wall")
camp = return_tower("camp")
spring_trap = return_tower("spring_trap")

lab = return_tower("lab")
wizard_tower = return_tower("wizard_tower")
air_defence = return_tower("air_defence")
inferno = return_tower("air_defence")

for tower in [lab, wizard_tower, air_defence, inferno]:
    tower.get_images()

# print(cannon.remaining_time(1, 5))

# print(spring_trap.remaining_time(1, 8))

# for tower in towers:
#     for level in tower.levels:
#         print(level.tower, level.level, level.gold)
