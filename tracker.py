import openpyxl as xl

tower_cells = []
level_cells = []

def get_cells():
    file = 'C:/Users/darre/OneDrive/Darren/clash_bot/tracker/tracker.xlsx'
    wb_values = xl.load_workbook(file, data_only=True)
    sheet_values = wb_values['1']

    for row in range(2, 31):
        tower = sheet_values.cell(row, 5).value
        if tower:
            tower = tower.lower().replace(" ", "_")
            tower_cells.append((tower, row))

    for column in range(6, sheet_values.max_column + 1):
        # print(sheet_values.max_column)
        level = sheet_values.cell(3, column).value
        level = column - 5
        # print(level, column)
        if level:
            level_cells.append((level, column))


# def find_all_values():
#     for tower, row in tower_cells:
#         for level, column in level_cells:
#             value = sheet_values.cell(row, column).value
#             if value and tower == "Archer Tower":
#                 print(f"{tower} Level:{level} Count:{value}")

def progress(account, tower_level):
    file = 'C:/Users/darre/OneDrive/Darren/clash_bot/tracker/tracker.xlsx'
    tower, level = tower_level
    wb = xl.load_workbook(file)
    sheet = wb[str(account.number)]
    row = next((x[1] for x in tower_cells if x[0] == tower), None)
    column = next((x[1] for x in level_cells if x[0] == level), None)
    # print(tower_cells)
    # print(level_cells)
    # print("Progress:", tower, row, column)
    # print("Value", sheet.cell(row, column).value)
    if row is None or column is None:
        print("Progress - row or column error", tower, level, row, column)
        return
    if sheet.cell(row, column).value and sheet.cell(row, column).value > 0:
        print("Progress success")
        sheet.cell(row, column).value = excel(sheet.cell(row, column).value) - 1
        sheet.cell(row, column + 1).value = excel(sheet.cell(row, column + 1).value) + 1
        wb.save(file)
        wb.close()

    print("Progress:", tower, row, column)

def excel(value):
    if not value: value = 0
    return value

def excel_write_tracker(account, tower_level, value):
    tower, level = tower_level
    file = 'C:/Users/darre/OneDrive/Darren/clash_bot/tracker/tracker.xlsx'
    wb = xl.load_workbook(file)
    sheet = wb[str(account.number)]
    row = next((x[1] for x in tower_cells if x[0] == tower), None)
    column = next((x[1] for x in level_cells if x[0] == level), None)
    print(tower_cells)
    print(level_cells)
    print("Excel write tracker:", tower, row, column)
    print("Value", sheet.cell(row, column).value)
    sheet.cell(row, column).value = value
    wb.save(file)
    wb.close()

def excel_read_tracker(account, tower_level):
    tower, level = tower_level
    file = 'C:/Users/darre/OneDrive/Darren/clash_bot/tracker/tracker.xlsx'
    wb = xl.load_workbook(file)
    sheet = wb[str(account.number)]
    row = next((x[1] for x in tower_cells if x[0] == tower), None)
    column = next((x[1] for x in level_cells if x[0] == level), None)
    value = sheet.cell(row, column).value
    print("Read tracker:", account, tower, level, value)
    return value



def excel_write(account_number, type, value):
    # print("Excel write", account_number, type, value)
    file = 'C:/Users/darre/OneDrive/Darren/clash_bot/tracker/info.xlsx'
    wb = xl.load_workbook(file)
    sheet = wb['Sheet1']
    if type == "next_completion":
        row, column = 2 + account_number, 2
        sheet.cell(row, column).value = value[0]
        sheet.cell(row, column + 1).value = value[1]
    elif type == "completion":
        row = sheet.max_row + 1
        sheet.cell(row, 1).value = account_number
        sheet.cell(row, 2).value = value[0]
        sheet.cell(row, 3).value = value[1]
    else:
        return
    wb.save(file)
    wb.close()

def excel_read(account_number, type):
    # print("Excel read", account_number, type)
    file = 'C:/Users/darre/OneDrive/Darren/clash_bot/tracker/info.xlsx'
    wb = xl.load_workbook(file)
    sheet = wb['Sheet1']
    if type == "next_completion":
        row, column = 2 + account_number, 2
        tower = sheet.cell(row, column).value
        level = sheet.cell(row, column + 1).value
        return tower, level
    else:
        return 0,0
    wb.close()

def excel_clear():
    for x in range(1,5):
        excel_write(x, "next_completion", ("", ""))

# excel_clear()

# excel_write(1,"next_completion", ("cannon", 6))
# excel_write(2,"next_completion", ("bomb", 6))
# excel_write(3,"next_completion", ("giant bomb", 8))
# excel_write(4,"next_completion", ("tesla", 12))



get_cells()
# print(tower_cells)
# print(level_cells)
# print("PRE")
# find_all_values()
# print()
# progress("Archer Tower", 17)
# print("POST")
# find_all_values()


# def decrease_value(tower, level)
